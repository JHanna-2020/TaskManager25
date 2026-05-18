"""
Task Manager Web App
A Flask-based web application for managing tasks (Firestore-backed)
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash
from datetime import datetime, timedelta
import os
from dotenv import load_dotenv
import firebase_admin
from firebase_admin import credentials, firestore
from apscheduler.schedulers.background import BackgroundScheduler
from discord_utils import send_discord_message
import pandas as pd
from openpyxl.styles import PatternFill
from flask import send_file
import io

# Load environment variables
load_dotenv()

# Initialize Flask app
app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-here')

# Initialize Firebase/Firestore
try:
    if not firebase_admin._apps:
        cred = credentials.Certificate(os.getenv("FIREBASE_CREDENTIALS_PATH", "firebase-credentials.json"))
        firebase_admin.initialize_app(cred)
    db = firestore.client()
    tasks_col = db.collection("tasks")
    print("[DEBUG] Firestore connection successful.")
except Exception as e:
    print(f"[ERROR] Firestore connection failed: {e}")
    db = None
    tasks_col = None

# --- Helper Functions ---
def decode_recurrence_days(bitmask):
    day_map = {1: "Mon", 2: "Tue", 4: "Wed", 8: "Thu", 16: "Fri", 32: "Sat", 64: "Sun"}
    weekdays = [name for bit, name in day_map.items() if bitmask and (bitmask & bit)]
    return ", ".join(weekdays)

def calculate_next_occurrence(current_due, recurrence_days):
    day_map = {1: 0, 2: 1, 4: 2, 8: 3, 16: 4, 32: 5, 64: 6}
    selected_days = [day_map[bit] for bit in day_map.keys() if recurrence_days & bit]

    if not selected_days:
        return None

    next_date = current_due + timedelta(days=1)

    for i in range(7):
        check_date = next_date + timedelta(days=i)
        if check_date.weekday() in selected_days:
            return check_date.replace(hour=current_due.hour, minute=current_due.minute, second=current_due.second)

    return next_date + timedelta(days=7)

def create_future_recurring_instances(name, course, start_dt, due_dt, recurrence_days, parent_task_id):
    if tasks_col is None:
        return
    current_due = due_dt
    duration = due_dt - start_dt

    for week in range(1, 13):
        next_occurrence = calculate_next_occurrence(current_due, recurrence_days)
        if next_occurrence:
            new_start = next_occurrence - duration
            due_str = next_occurrence.strftime("%Y-%m-%d %H:%M:%S")

            results = list(tasks_col.where("name", "==", name).where("due", "==", due_str).limit(1).stream())
            if not results:
                try:
                    tasks_col.add({
                        "name": name,
                        "course": course,
                        "start": new_start.strftime("%Y-%m-%d %H:%M:%S"),
                        "due": due_str,
                        "status": "Not Started",
                        "recurrence_days": recurrence_days,
                        "parent_task_id": str(parent_task_id) if parent_task_id else None,
                        "is_recurring_instance": True
                    })
                except Exception as e:
                    print(f"Failed to create future recurring instance: {e}")

            current_due = next_occurrence
        else:
            break

def create_due_weekday_instance(name, course, start_dt, due_dt, parent_task_id):
    if tasks_col is None:
        return

    duration = due_dt - start_dt
    next_due = due_dt + timedelta(days=7)
    next_start = next_due - duration
    due_str = next_due.strftime("%Y-%m-%d %H:%M:%S")

    results = list(tasks_col.where("name", "==", name).where("due", "==", due_str).limit(1).stream())
    if not results:
        try:
            tasks_col.add({
                "name": name,
                "course": course,
                "start": next_start.strftime("%Y-%m-%d %H:%M:%S"),
                "due": due_str,
                "status": "Not Started",
                "recurrence_days": -1,
                "parent_task_id": str(parent_task_id) if parent_task_id else None,
                "is_recurring_instance": True
            })
        except Exception as e:
            print(f"Failed to create due weekday instance: {e}")

# --- Routes ---
@app.route('/')
def index():
    if tasks_col is None:
        flash("Database connection error", "error")
        return render_template('index.html', tasks=[], total_count=0, active_count=0, completed_count=0)

    tasks = []
    active_tasks = []
    completed_tasks = []

    try:
        all_docs = tasks_col.stream()

        for doc in all_docs:
            task = doc.to_dict()
            task['id'] = doc.id

            try:
                start_dt = datetime.strptime(task.get("start", ""), "%Y-%m-%d %H:%M:%S")
                due_dt = datetime.strptime(task.get("due", ""), "%Y-%m-%d %H:%M:%S")
                task['start_formatted'] = start_dt.strftime("%m/%d/%y %I:%M %p")
                task['due_formatted'] = due_dt.strftime("%m/%d/%y %I:%M %p")
            except (ValueError, TypeError):
                task['start_formatted'] = "Invalid Date"
                task['due_formatted'] = "Invalid Date"

            if task.get("status") in ["Completed", "Graded"]:
                completed_tasks.append(task)
            else:
                active_tasks.append(task)

        active_tasks.sort(key=lambda x: x.get("due", ""))
        completed_tasks.sort(key=lambda x: x.get("due", ""))
        tasks = active_tasks + completed_tasks

    except Exception as e:
        flash(f"Error loading tasks: {e}", "error")
        print(f"[ERROR] Exception in index(): {e}")

    return render_template('index.html', tasks=tasks, total_count=len(tasks), active_count=len(active_tasks), completed_count=len(completed_tasks))

@app.route('/add_task', methods=['GET', 'POST'])
def add_task():
    if tasks_col is None:
        flash("Database connection error", "error")
        return render_template('add_task.html')

    if request.method == 'POST':
        try:
            name = request.form.get('name')
            course = request.form.get('course')
            start_date = request.form.get('start_date')
            start_time = request.form.get('start_time')
            due_date = request.form.get('due_date')
            due_time = request.form.get('due_time')
            status = request.form.get('status')
            reminder_hours = int(request.form.get('reminder_hours', 24))

            if not all([name, course, start_date, start_time, due_date, due_time]):
                flash('All fields are required!', 'error')
                return render_template('add_task.html')

            start_dt = datetime.strptime(f"{start_date} {start_time}", "%Y-%m-%d %H:%M")
            due_dt = datetime.strptime(f"{due_date} {due_time}", "%Y-%m-%d %H:%M")

            recurrence_type = request.form.get('recurrence_type', 'none')
            recurrence_days = 0

            if recurrence_type == 'weekly':
                day_map = {'mon': 1, 'tue': 2, 'wed': 4, 'thu': 8, 'fri': 16, 'sat': 32, 'sun': 64}
                for day in day_map.keys():
                    if request.form.get(f'recurrence_{day}'):
                        recurrence_days |= day_map[day]
            elif recurrence_type == 'due_weekday':
                recurrence_days = -1

            _, doc_ref = tasks_col.add({
                "name": name,
                "course": course,
                "start": start_dt.strftime("%Y-%m-%d %H:%M:%S"),
                "due": due_dt.strftime("%Y-%m-%d %H:%M:%S"),
                "status": status,
                "recurrence_days": recurrence_days,
                "is_recurring_instance": False,
                "reminder_hours": reminder_hours,
                "reminder_sent": 0
            })

            if recurrence_days > 0:
                create_future_recurring_instances(name, course, start_dt, due_dt, recurrence_days, doc_ref.id)
            elif recurrence_days == -1:
                create_due_weekday_instance(name, course, start_dt, due_dt, doc_ref.id)

            flash('Task added successfully!', 'success')
            return redirect(url_for('index'))

        except ValueError as e:
            flash(f'Invalid date/time format: {e}', 'error')
        except Exception as e:
            flash(f'Error adding task: {e}', 'error')

    return render_template('add_task.html')

@app.route('/edit_task/<task_id>', methods=['GET', 'POST'])
def edit_task(task_id):
    if tasks_col is None:
        flash("Database connection error", "error")
        return redirect(url_for('index'))

    if request.method == 'POST':
        try:
            name = request.form.get('name')
            course = request.form.get('course')
            start_date = request.form.get('start_date')
            start_time = request.form.get('start_time')
            due_date = request.form.get('due_date')
            due_time = request.form.get('due_time')
            status = request.form.get('status')

            if not all([name, course, start_date, start_time, due_date, due_time]):
                flash('All fields are required!', 'error')
                return redirect(url_for('edit_task', task_id=task_id))

            start_dt = datetime.strptime(f"{start_date} {start_time}", "%Y-%m-%d %H:%M")
            due_dt = datetime.strptime(f"{due_date} {due_time}", "%Y-%m-%d %H:%M")

            tasks_col.document(task_id).update({
                "name": name,
                "course": course,
                "start": start_dt.strftime("%Y-%m-%d %H:%M:%S"),
                "due": due_dt.strftime("%Y-%m-%d %H:%M:%S"),
                "status": status
            })

            flash('Task updated successfully!', 'success')
            return redirect(url_for('index'))

        except ValueError as e:
            flash(f'Invalid date/time format: {e}', 'error')
        except Exception as e:
            flash(f'Error updating task: {e}', 'error')

    try:
        doc = tasks_col.document(task_id).get()
        if doc.exists:
            task = doc.to_dict()
            task['id'] = doc.id
            start_dt = datetime.strptime(task.get("start"), "%Y-%m-%d %H:%M:%S")
            due_dt = datetime.strptime(task.get("due"), "%Y-%m-%d %H:%M:%S")
            task['start_date'] = start_dt.strftime("%Y-%m-%d")
            task['start_time'] = start_dt.strftime("%H:%M")
            task['due_date'] = due_dt.strftime("%Y-%m-%d")
            task['due_time'] = due_dt.strftime("%H:%M")
            return render_template('edit_task.html', task=task)
        else:
            flash('Task not found!', 'error')
            return redirect(url_for('index'))
    except Exception as e:
        flash(f'Error loading task: {e}', 'error')
        return redirect(url_for('index'))


@app.route('/delete_task/<task_id>')
def delete_task(task_id):
    if tasks_col is None:
        flash("Database connection error", "error")
        return redirect(url_for('index'))

    try:
        tasks_col.document(task_id).delete()
        flash('Task deleted successfully!', 'success')
    except Exception as e:
        flash(f'Error deleting task: {e}', 'error')

    return redirect(url_for('index'))

@app.route('/update_status/<task_id>/<status>')
def update_status(task_id, status):
    if tasks_col is None:
        flash("Database connection error", "error")
        return redirect(url_for('index'))

    try:
        tasks_col.document(task_id).update({"status": status})
        flash('Status updated successfully!', 'success')
    except Exception as e:
        flash(f'Error updating status: {e}', 'error')

    return redirect(url_for('index'))

@app.route('/view_by_class/<class_name>')
def view_by_class(class_name):
    if tasks_col is None:
        flash("Database connection error", "error")
        return render_template('view_by_class.html', tasks=[], class_name=class_name, class_count=0)

    tasks = []
    try:
        docs = tasks_col.where("course", "==", class_name).stream()
        for doc in docs:
            task = doc.to_dict()
            task['id'] = doc.id
            try:
                start_dt = datetime.strptime(task.get("start", ""), "%Y-%m-%d %H:%M:%S")
                due_dt = datetime.strptime(task.get("due", ""), "%Y-%m-%d %H:%M:%S")
                task['start_formatted'] = start_dt.strftime("%m/%d/%y %I:%M %p")
                task['due_formatted'] = due_dt.strftime("%m/%d/%y %I:%M %p")
            except (ValueError, TypeError):
                task['start_formatted'] = "Invalid Date"
                task['due_formatted'] = "Invalid Date"
            tasks.append(task)
        tasks.sort(key=lambda x: x.get("due", ""))
    except Exception as e:
        flash(f"Error loading tasks: {e}", "error")

    return render_template('view_by_class.html', tasks=tasks, class_name=class_name, class_count=len(tasks))

@app.route('/view_completed')
def view_completed():
    if tasks_col is None:
        flash("Database connection error", "error")
        return render_template('view_completed.html', tasks=[], completed_count=0)

    tasks = []
    try:
        docs = tasks_col.where("status", "in", ["Completed", "Graded"]).stream()
        for doc in docs:
            task = doc.to_dict()
            task['id'] = doc.id
            try:
                start_dt = datetime.strptime(task.get("start", ""), "%Y-%m-%d %H:%M:%S")
                due_dt = datetime.strptime(task.get("due", ""), "%Y-%m-%d %H:%M:%S")
                task['start_formatted'] = start_dt.strftime("%m/%d/%y %I:%M %p")
                task['due_formatted'] = due_dt.strftime("%m/%d/%y %I:%M %p")
            except (ValueError, TypeError):
                task['start_formatted'] = "Invalid Date"
                task['due_formatted'] = "Invalid Date"
            tasks.append(task)
        tasks.sort(key=lambda x: x.get("due", ""))
    except Exception as e:
        flash(f"Error loading completed tasks: {e}", "error")

    return render_template('view_completed.html', tasks=tasks, completed_count=len(tasks))

@app.route('/view_active')
def view_active():
    if tasks_col is None:
        flash("Database connection error", "error")
        return render_template('view_active.html', tasks=[], active_count=0)

    tasks = []
    try:
        docs = tasks_col.where("status", "in", ["Not Started", "In Progress"]).stream()
        for doc in docs:
            task = doc.to_dict()
            task['id'] = doc.id
            try:
                start_dt = datetime.strptime(task.get("start", ""), "%Y-%m-%d %H:%M:%S")
                due_dt = datetime.strptime(task.get("due", ""), "%Y-%m-%d %H:%M:%S")
                task['start_formatted'] = start_dt.strftime("%m/%d/%y %I:%M %p")
                task['due_formatted'] = due_dt.strftime("%m/%d/%y %I:%M %p")
            except (ValueError, TypeError):
                task['start_formatted'] = "Invalid Date"
                task['due_formatted'] = "Invalid Date"
            tasks.append(task)
        tasks.sort(key=lambda x: x.get("due", ""))
    except Exception as e:
        flash(f"Error loading active tasks: {e}", "error")

    return render_template('view_active.html', tasks=tasks, active_count=len(tasks))

@app.route('/delete_all_tasks')
def delete_all_tasks():
    if tasks_col is None:
        flash("Database connection error", "error")
        return redirect(url_for('index'))

    try:
        docs = list(tasks_col.stream())
        for doc in docs:
            doc.reference.delete()
        flash(f'Successfully deleted {len(docs)} tasks!', 'success')
    except Exception as e:
        flash(f'Error deleting all tasks: {e}', 'error')
    return redirect(url_for('index'))

def check_reminders():
    with app.app_context():
        if tasks_col is None:
            return
        now = datetime.now()
        docs = tasks_col.where("reminder_sent", "==", 0).stream()

        for doc in docs:
            task = doc.to_dict()
            if task.get("status") == "Completed":
                continue
            due = datetime.strptime(task["due"], "%Y-%m-%d %H:%M:%S")
            reminder_hours = task.get("reminder_hours", 24)
            reminder_time = due - timedelta(hours=reminder_hours)

            if reminder_time <= now < due:
                message = f"Reminder: Your task '{task['name']}' is due at {due.strftime('%I:%M %p')}."
                try:
                    send_discord_message(message)
                    tasks_col.document(doc.id).update({"reminder_sent": 1})
                except Exception as e:
                    print(f"Failed to send Discord reminder: {e}")

scheduler = BackgroundScheduler(daemon=True)
scheduler.add_job(check_reminders, 'interval', seconds=60)
scheduler.start()

@app.route('/export')
def export_to_excel():
    if tasks_col is None:
        flash("Database connection error", "error")
        return redirect(url_for('index'))

    docs = list(tasks_col.stream())
    if not docs:
        flash('No tasks to export!', 'info')
        return redirect(url_for('index'))

    tasks = [doc.to_dict() for doc in docs]
    df = pd.DataFrame(tasks)

    df['start'] = pd.to_datetime(df['start']).dt.strftime('%m/%d/%y %I:%M %p')
    df['due'] = pd.to_datetime(df['due']).dt.strftime('%m/%d/%y %I:%M %p')

    df = df[['name', 'course', 'start', 'due', 'status']]
    df.columns = ['Task Name', 'Class', 'Start Date', 'Due Date', 'Status']

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Tasks')

    workbook = writer.book
    worksheet = writer.sheets['Tasks']

    colors = {
        'Not Started': 'FF7171',
        'In Progress': 'FFFACD',
        'Completed': 'D0F0C0',
        'Graded': 'ADD8E6'
    }

    for index, row in df.iterrows():
        status = row['Status']
        if status in colors:
            fill = PatternFill(start_color=colors[status], end_color=colors[status], fill_type="solid")
            for col_idx in range(1, len(df.columns) + 1):
                worksheet.cell(row=index + 2, column=col_idx).fill = fill

    writer.close()
    output.seek(0)

    current_date = datetime.now().strftime("%m-%d-%y")
    filename = f"tasks_{current_date}.xlsx"

    return send_file(output, download_name=filename, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=8080)
